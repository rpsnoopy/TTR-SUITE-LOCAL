"""
TTR-SUITE — Consolidated Results Builder
=========================================
Merges all per-run CSVs into a single comparison XLSX.

Sheets produced:
  1. Accuracy          — score per model × benchmark
  2. Throughput        — tok/s per model
  3. LegalBench        — per category (6 cat)
  4. CUAD              — per category (8 cat)
  5. TTR-Score         — weighted composite score (TTR relevance)
  6. Micro-heatmap     — all categories in one colour-coded view
  7. Rankings          — rank per micro-category
  8. TTR-Radar         — 5 synthetic axes for TTR use-case
  9. Partial runs      — incomplete models

Usage:
    python consolidate_results.py [--results-dir PATH] [--output PATH]

Re-run whenever a new model is added (gpt-4o, deepcoder, etc.).
Multiple runs of the same model are averaged automatically.
"""

import argparse
import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── Config ─────────────────────────────────────────────────────────────────────

RESULTS_DIR       = Path("C:/TTR_Benchmark/results")
BENCHMARK_RESULTS = Path(__file__).parent / "results"
BENCHMARKS        = ["legalbench", "cuad", "ifeval", "mmlupro"]

# Display order — add new models here before running
MODEL_ORDER = [
    "claude-sonnet-4-6",
    "claude-sonnet-4-5",
    "gpt-oss-20b",
    "gpt-4o",
    "gpt-4o-mini",
    "deepcoder-14b",
    "qwen2.5-coder-14b",
    "mistral-small-24b",
    "qwen3-30b-a3b",
    "qwen3-14b",
    "qwen3-32b",
]

# ── TTR relevance weights ───────────────────────────────────────────────────────
# Reflects how much each sub-task matters for IP/contract analysis at TTR.
# Scale 0.5–3: 3=critical, 2=high, 1=useful, 0.5=informative only.

TTR_WEIGHTS = {
    # LegalBench categories
    "lb:issue-spotting":          3.0,   # identify IP issues in contract text
    "lb:rule-application":        3.0,   # apply rules to specific contract facts
    "lb:interpretation":          3.0,   # interpret ambiguous clauses
    "lb:rule-conclusion":         2.0,   # reach correct legal conclusion
    "lb:rule-recall":             1.0,   # less critical with RAG
    "lb:rhetorical-understanding":0.5,   # least relevant for TTR
    # Benchmark totals
    "cuad":                       3.0,   # direct contract clause extraction
    "ifeval":                     2.0,   # structured/instruction-following output
    "mmlupro":                    1.0,   # legal knowledge breadth
}

# ── 5 synthetic radar axes ──────────────────────────────────────────────────────
RADAR_AXES = {
    "Contract\nExtraction":   [("cuad", None)],
    "Legal\nReasoning":       [("legalbench", "rule-application"),
                               ("legalbench", "rule-conclusion"),
                               ("legalbench", "interpretation")],
    "Issue\nSpotting":        [("legalbench", "issue-spotting")],
    "Instruction\nFollowing": [("ifeval", None)],
    "Legal\nKnowledge":       [("mmlupro", None)],
}

# ── Colour helpers ──────────────────────────────────────────────────────────────

_HDR_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
_HDR_FONT   = Font(color="FFFFFF", bold=True)
_SUBHDR_FILL= PatternFill("solid", fgColor="2E75B6")   # medium blue
_BEST_FONT  = Font(bold=True, color="1F4E79")
_THIN       = Side(style="thin", color="CCCCCC")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header_row(ws, row_num: int, ncols: int, fill=_HDR_FILL):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _BORDER


def _apply_color_scale(ws, min_row, max_row, min_col, max_col):
    """Green=high, red=low, 3-color scale."""
    rule = ColorScaleRule(
        start_type="min",  start_color="F8696B",   # red
        mid_type="num",    mid_value=50, mid_color="FFEB84",  # yellow at 50%
        end_type="max",    end_color="63BE7B",     # green
    )
    start = f"{get_column_letter(min_col)}{min_row}"
    end   = f"{get_column_letter(max_col)}{max_row}"
    ws.conditional_formatting.add(f"{start}:{end}", rule)


def _bold_max_per_col(ws, data_row_start, data_row_end, col_start, col_end):
    """Bold the best (max) value in each column."""
    for col in range(col_start, col_end + 1):
        best_val, best_row = None, None
        for row in range(data_row_start, data_row_end + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None and isinstance(v, (int, float)):
                if best_val is None or v > best_val:
                    best_val, best_row = v, row
        if best_row:
            ws.cell(row=best_row, column=col).font = _BEST_FONT


# ── Load & merge ────────────────────────────────────────────────────────────────

def load_all_csvs(results_dir: Path) -> pd.DataFrame:
    csv_files = sorted(results_dir.glob("raw_results_*.csv"))
    if not csv_files:
        sys.exit(f"No raw_results_*.csv found in {results_dir}")

    frames = []
    for f in csv_files:
        df = pd.read_csv(f, low_memory=False)
        if df.empty or "model" not in df.columns:
            continue
        df["_source_file"] = f.name
        frames.append(df)
        models = sorted(df["model"].dropna().unique())
        print(f"  {f.name}: {len(df)} rows — {models}")

    df_all = pd.concat(frames, ignore_index=True)

    # Deduplicate: if the same (model, benchmark, task_id) appears in multiple
    # runs, keep only the LAST entry (CSVs are sorted by filename = timestamp,
    # so the latest run wins).  This prevents old buggy runs from contaminating
    # re-run benchmarks.
    if "task_id" in df_all.columns:
        before = len(df_all)
        df_all = df_all.drop_duplicates(
            subset=["model", "benchmark", "task_id"], keep="last"
        )
        dropped = before - len(df_all)
        if dropped:
            print(f"  [dedup] dropped {dropped} duplicate rows (kept latest run per task)")

    return df_all


def classify_models(df: pd.DataFrame):
    complete, partial = [], []
    for model, g in df.groupby("model"):
        done = set(g["benchmark"].dropna().unique())
        if set(BENCHMARKS).issubset(done):
            complete.append(model)
        else:
            partial.append((model, sorted(done)))
    return complete, partial


def _sort_models(df: pd.DataFrame) -> pd.DataFrame:
    order = {m: i for i, m in enumerate(MODEL_ORDER)}
    idx = df.index.to_series().map(lambda m: order.get(m, 9000 + sum(ord(c) for c in m)))
    return df.loc[idx.sort_values().index]


# ── Sheet 1: Accuracy ──────────────────────────────────────────────────────────

def sheet_accuracy(df: pd.DataFrame, writer: pd.ExcelWriter) -> None:
    pivot = (
        df.groupby(["model", "benchmark"])["is_correct"]
        .mean().mul(100).round(1)
        .unstack("benchmark")
        .reindex(columns=BENCHMARKS)
    )
    pivot.columns.name = None
    pivot.index.name   = "Model"
    pivot = _sort_models(pivot)
    pivot["AVG"] = pivot[BENCHMARKS].mean(axis=1).round(1)
    pivot.to_excel(writer, sheet_name="Accuracy")

    ws = writer.sheets["Accuracy"]
    _style_header_row(ws, 1, len(pivot.columns) + 1)
    _apply_color_scale(ws, 2, len(pivot) + 1, 2, len(pivot.columns) + 1)
    _bold_max_per_col(ws, 2, len(pivot) + 1, 2, len(pivot.columns) + 1)
    for col in range(1, len(pivot.columns) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 14


# ── Sheet 2: Throughput ────────────────────────────────────────────────────────

def sheet_throughput(df: pd.DataFrame, writer: pd.ExcelWriter) -> None:
    overall      = df.groupby("model")["tok_s"].mean().round(1).rename("avg tok/s")
    thinking     = (df[df["thinking_tokens"] > 0].groupby("model")["tok_s"]
                    .mean().round(1).rename("thinking tok/s"))
    non_thinking = (df[df["thinking_tokens"] == 0].groupby("model")["tok_s"]
                    .mean().round(1).rename("non-thinking tok/s"))
    out = pd.concat([overall, thinking, non_thinking], axis=1)
    out.index.name = "Model"
    out = _sort_models(out)
    out.to_excel(writer, sheet_name="Throughput")

    ws = writer.sheets["Throughput"]
    _style_header_row(ws, 1, len(out.columns) + 1)
    _apply_color_scale(ws, 2, len(out) + 1, 2, 2)   # colour only avg tok/s
    _bold_max_per_col(ws, 2, len(out) + 1, 2, 2)
    for col in range(1, len(out.columns) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 18


# ── Sheet 3: LegalBench categories ────────────────────────────────────────────

LB_CAT_ORDER = [
    "issue-spotting", "rule-application", "interpretation",
    "rule-conclusion", "rule-recall", "rhetorical-understanding",
]
LB_CAT_LABELS = {
    "issue-spotting":          "Issue Spotting",
    "rule-application":        "Rule Application",
    "interpretation":          "Interpretation",
    "rule-conclusion":         "Rule Conclusion",
    "rule-recall":             "Rule Recall",
    "rhetorical-understanding":"Rhetorical Understanding",
}

def sheet_legalbench(df: pd.DataFrame, writer: pd.ExcelWriter) -> None:
    lb = df[df["benchmark"] == "legalbench"]
    if lb.empty:
        return
    pivot = (
        lb.groupby(["model", "category"])["is_correct"]
        .mean().mul(100).round(1)
        .unstack("category")
        .reindex(columns=LB_CAT_ORDER)
    )
    pivot.rename(columns=LB_CAT_LABELS, inplace=True)
    pivot.columns.name = None
    pivot.index.name   = "Model"
    pivot = _sort_models(pivot)
    pivot["AVG"] = pivot.mean(axis=1).round(1)
    pivot.to_excel(writer, sheet_name="LegalBench")

    ws = writer.sheets["LegalBench"]
    _style_header_row(ws, 1, len(pivot.columns) + 1)
    _apply_color_scale(ws, 2, len(pivot) + 1, 2, len(pivot.columns) + 1)
    _bold_max_per_col(ws, 2, len(pivot) + 1, 2, len(pivot.columns) + 1)
    for col in range(1, len(pivot.columns) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 22


# ── Sheet 4: CUAD categories ──────────────────────────────────────────────────

CUAD_CAT_ORDER = [
    "Change-of-Control", "Anti-Assignment", "Non-Compete", "Exclusivity",
    "Governing-Law", "Renewal-Term", "Expiration-Date", "Parties",
]

def sheet_cuad(df: pd.DataFrame, writer: pd.ExcelWriter) -> None:
    cuad = df[df["benchmark"] == "cuad"]
    if cuad.empty:
        return
    existing_cats = [c for c in CUAD_CAT_ORDER if c in cuad["category"].unique()]
    pivot = (
        cuad.groupby(["model", "category"])["is_correct"]
        .mean().mul(100).round(1)
        .unstack("category")
        .reindex(columns=existing_cats)
    )
    pivot.columns.name = None
    pivot.index.name   = "Model"
    pivot = _sort_models(pivot)
    pivot["AVG"] = pivot.mean(axis=1).round(1)
    pivot.to_excel(writer, sheet_name="CUAD")

    ws = writer.sheets["CUAD"]
    _style_header_row(ws, 1, len(pivot.columns) + 1)
    _apply_color_scale(ws, 2, len(pivot) + 1, 2, len(pivot.columns) + 1)
    _bold_max_per_col(ws, 2, len(pivot) + 1, 2, len(pivot.columns) + 1)
    for col in range(1, len(pivot.columns) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 18


# ── Sheet 5: TTR-Score ────────────────────────────────────────────────────────

def _ttr_score(df: pd.DataFrame) -> pd.Series:
    """Compute weighted TTR score (0-100) per model."""
    scores = {}
    for model, g in df.groupby("model"):
        total_w, weighted_sum = 0.0, 0.0

        # LegalBench sub-categories
        lb = g[g["benchmark"] == "legalbench"]
        for cat in LB_CAT_ORDER:
            key = f"lb:{cat}"
            w = TTR_WEIGHTS.get(key, 1.0)
            sub = lb[lb["category"] == cat]
            if not sub.empty:
                weighted_sum += sub["is_correct"].mean() * 100 * w
                total_w += w

        # CUAD, IFEval, MMLU-Pro as totals
        for bm in ["cuad", "ifeval", "mmlupro"]:
            w = TTR_WEIGHTS.get(bm, 1.0)
            sub = g[g["benchmark"] == bm]
            if not sub.empty:
                weighted_sum += sub["is_correct"].mean() * 100 * w
                total_w += w

        scores[model] = round(weighted_sum / total_w, 1) if total_w > 0 else float("nan")

    return pd.Series(scores, name="TTR-Score (weighted)")


def sheet_ttr_score(df: pd.DataFrame, writer: pd.ExcelWriter) -> None:
    ttr = _ttr_score(df).to_frame()
    ttr.index.name = "Model"
    ttr = _sort_models(ttr)

    # Add component scores for context
    for bm, label in [("legalbench","LegalBench"), ("cuad","CUAD"),
                       ("ifeval","IFEval"), ("mmlupro","MMLU-Pro")]:
        ttr[label] = (
            df.groupby(["model","benchmark"])["is_correct"]
            .mean().mul(100).round(1)
            .unstack("benchmark")
            .get(bm, pd.Series(dtype=float))
            .reindex(ttr.index)
        )

    # Weight annotations
    weight_row = {"TTR-Score (weighted)": "COMPOSITE",
                  "LegalBench": f"×{TTR_WEIGHTS['lb:rule-application']}–{TTR_WEIGHTS['lb:rhetorical-understanding']}",
                  "CUAD": f"×{TTR_WEIGHTS['cuad']}",
                  "IFEval": f"×{TTR_WEIGHTS['ifeval']}",
                  "MMLU-Pro": f"×{TTR_WEIGHTS['mmlupro']}"}

    ttr.to_excel(writer, sheet_name="TTR-Score")
    ws = writer.sheets["TTR-Score"]
    _style_header_row(ws, 1, len(ttr.columns) + 1)
    _apply_color_scale(ws, 2, len(ttr) + 1, 2, 2)   # colour TTR-Score col only
    _bold_max_per_col(ws, 2, len(ttr) + 1, 2, 2)

    # Write weight annotation row
    ann_row = len(ttr) + 3
    ws.cell(row=ann_row, column=1, value="Weights →")
    ws.cell(row=ann_row, column=1).font = Font(italic=True, color="666666")
    for ci, col in enumerate(ttr.columns, start=2):
        ws.cell(row=ann_row, column=ci, value=weight_row.get(col, ""))
        ws.cell(row=ann_row, column=ci).font = Font(italic=True, color="666666")
        ws.cell(row=ann_row, column=ci).alignment = Alignment(horizontal="center")

    for col in range(1, len(ttr.columns) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 20


# ── Sheet 6: Micro-heatmap ────────────────────────────────────────────────────

def sheet_micro_heatmap(df: pd.DataFrame, writer: pd.ExcelWriter) -> None:
    """All benchmark sub-categories + benchmark totals in one coloured table."""
    rows = {}
    for model, g in df.groupby("model"):
        row = {}
        # LegalBench categories
        lb = g[g["benchmark"] == "legalbench"]
        for cat in LB_CAT_ORDER:
            sub = lb[lb["category"] == cat]
            row[f"LB: {LB_CAT_LABELS.get(cat, cat)}"] = (
                round(sub["is_correct"].mean() * 100, 1) if not sub.empty else None
            )
        row["LB AVG"] = round(lb["is_correct"].mean() * 100, 1) if not lb.empty else None

        # CUAD categories
        cuad = g[g["benchmark"] == "cuad"]
        for cat in CUAD_CAT_ORDER:
            sub = cuad[cuad["category"] == cat]
            row[f"CUAD: {cat}"] = (
                round(sub["is_correct"].mean() * 100, 1) if not sub.empty else None
            )
        row["CUAD AVG"] = round(cuad["is_correct"].mean() * 100, 1) if not cuad.empty else None

        # IFEval and MMLU-Pro totals
        for bm, label in [("ifeval", "IFEval"), ("mmlupro", "MMLU-Pro")]:
            sub = g[g["benchmark"] == bm]
            row[label] = round(sub["is_correct"].mean() * 100, 1) if not sub.empty else None

        rows[model] = row

    heatmap = pd.DataFrame(rows).T
    heatmap.index.name = "Model"
    heatmap = _sort_models(heatmap)
    heatmap.to_excel(writer, sheet_name="Micro-heatmap")

    ws = writer.sheets["Micro-heatmap"]
    ncols = len(heatmap.columns) + 1
    _style_header_row(ws, 1, ncols)

    # Section sub-headers (LB / CUAD / other)
    lb_cols   = [i+2 for i, c in enumerate(heatmap.columns) if c.startswith("LB")]
    cuad_cols = [i+2 for i, c in enumerate(heatmap.columns) if c.startswith("CUAD")]

    for col_idx in lb_cols:
        ws.cell(row=1, column=col_idx).fill = PatternFill("solid", fgColor="2E4057")
    for col_idx in cuad_cols:
        ws.cell(row=1, column=col_idx).fill = PatternFill("solid", fgColor="1B4332")

    _apply_color_scale(ws, 2, len(heatmap) + 1, 2, ncols)
    _bold_max_per_col(ws, 2, len(heatmap) + 1, 2, ncols)
    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.freeze_panes = "B2"


# ── Sheet 7: Rankings ─────────────────────────────────────────────────────────

def sheet_rankings(df: pd.DataFrame, writer: pd.ExcelWriter) -> None:
    """Rank models (1=best) for each micro-category."""
    rows = {}
    for model, g in df.groupby("model"):
        row = {}
        lb = g[g["benchmark"] == "legalbench"]
        for cat in LB_CAT_ORDER:
            sub = lb[lb["category"] == cat]
            row[f"LB: {LB_CAT_LABELS.get(cat, cat)}"] = (
                round(sub["is_correct"].mean() * 100, 1) if not sub.empty else None
            )
        cuad = g[g["benchmark"] == "cuad"]
        for cat in CUAD_CAT_ORDER:
            sub = cuad[cuad["category"] == cat]
            row[f"CUAD: {cat}"] = (
                round(sub["is_correct"].mean() * 100, 1) if not sub.empty else None
            )
        for bm, label in [("ifeval", "IFEval"), ("mmlupro", "MMLU-Pro"), ("cuad", "CUAD AVG"), ("legalbench", "LB AVG")]:
            sub = g[g["benchmark"] == bm]
            row[label] = round(sub["is_correct"].mean() * 100, 1) if not sub.empty else None
        rows[model] = row

    scores = pd.DataFrame(rows).T
    ranks  = scores.rank(ascending=False, method="min").astype("Int64")
    ranks.index.name = "Model"
    ranks = _sort_models(ranks)
    ranks.to_excel(writer, sheet_name="Rankings")

    ws = writer.sheets["Rankings"]
    _style_header_row(ws, 1, len(ranks.columns) + 1)

    gold   = PatternFill("solid", fgColor="FFD700")
    silver = PatternFill("solid", fgColor="C0C0C0")
    bronze = PatternFill("solid", fgColor="CD7F32")

    for row in range(2, len(ranks) + 2):
        for col in range(2, len(ranks.columns) + 2):
            cell = ws.cell(row=row, column=col)
            if cell.value == 1:
                cell.fill = gold
                cell.font = Font(bold=True)
            elif cell.value == 2:
                cell.fill = silver
            elif cell.value == 3:
                cell.fill = bronze

    for col in range(1, len(ranks.columns) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.freeze_panes = "B2"


# ── Sheet 8: TTR-Radar ────────────────────────────────────────────────────────

def sheet_ttr_radar(df: pd.DataFrame, writer: pd.ExcelWriter) -> None:
    """5-axis radar data table for TTR use-case."""
    axes = list(RADAR_AXES.keys())
    rows = {}
    for model, g in df.groupby("model"):
        row = {}
        for axis_name, components in RADAR_AXES.items():
            vals = []
            for bm, cat in components:
                sub = g[g["benchmark"] == bm]
                if cat:
                    sub = sub[sub["category"] == cat]
                if not sub.empty:
                    vals.append(sub["is_correct"].mean() * 100)
            row[axis_name] = round(sum(vals) / len(vals), 1) if vals else None
        rows[model] = row

    radar = pd.DataFrame(rows).T
    radar.index.name = "Model"
    radar = _sort_models(radar)

    # Add TTR-Score
    radar.insert(0, "TTR-Score", _ttr_score(df).reindex(radar.index))

    radar.to_excel(writer, sheet_name="TTR-Radar")
    ws = writer.sheets["TTR-Radar"]
    _style_header_row(ws, 1, len(radar.columns) + 1)
    _apply_color_scale(ws, 2, len(radar) + 1, 2, len(radar.columns) + 1)
    _bold_max_per_col(ws, 2, len(radar) + 1, 2, len(radar.columns) + 1)

    # Axis descriptions
    descs = {
        "TTR-Score":            "Weighted composite (TTR relevance)",
        "Contract\nExtraction": "CUAD avg F1 — clause extraction accuracy",
        "Legal\nReasoning":     "LB: Rule-Application + Conclusion + Interpretation avg",
        "Issue\nSpotting":      "LB: Issue-Spotting category",
        "Instruction\nFollowing": "IFEval — structured output compliance",
        "Legal\nKnowledge":     "MMLU-Pro Law — background legal knowledge",
    }
    ann_row = len(radar) + 3
    ws.cell(row=ann_row, column=1, value="Description →").font = Font(italic=True, color="666666")
    for ci, col in enumerate(radar.columns, start=2):
        ws.cell(row=ann_row, column=ci, value=descs.get(col, "")).font = Font(italic=True, color="666666")
        ws.cell(row=ann_row, column=ci).alignment = Alignment(wrap_text=True)

    for col in range(1, len(radar.columns) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "B2"


# ── Sheet 9: Partial runs ─────────────────────────────────────────────────────

def sheet_partial(partial: list, writer: pd.ExcelWriter) -> None:
    if not partial:
        return
    rows = [{"Model": m, "Benchmarks completed": ", ".join(b),
             "Missing": ", ".join(set(BENCHMARKS) - set(b))} for m, b in partial]
    pd.DataFrame(rows).to_excel(writer, sheet_name="Partial runs", index=False)
    ws = writer.sheets["Partial runs"]
    _style_header_row(ws, 1, 3)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Build consolidated benchmark XLSX")
    parser.add_argument("--output", type=Path,
                        default=BENCHMARK_RESULTS / f"consolidated_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    parser.add_argument("--results-dir", type=Path, default=RESULTS_DIR)
    args = parser.parse_args()

    print(f"\nLoading CSVs from {args.results_dir} ...")
    df = load_all_csvs(args.results_dir)
    print(f"Total rows: {len(df)}")

    complete, partial = classify_models(df)
    print(f"Complete models ({len(complete)}): {complete}")
    if partial:
        print(f"Partial  models ({len(partial)}): {[m for m,_ in partial]}")

    df_c = df[df["model"].isin(complete)]

    args.output.parent.mkdir(parents=True, exist_ok=True)
    print(f"\nWriting {args.output} ...")
    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        sheet_accuracy(df_c, writer)
        sheet_throughput(df_c, writer)
        sheet_legalbench(df_c, writer)
        sheet_cuad(df_c, writer)
        sheet_ttr_score(df_c, writer)
        sheet_micro_heatmap(df_c, writer)
        sheet_rankings(df_c, writer)
        sheet_ttr_radar(df_c, writer)
        sheet_partial(partial, writer)

    print(f"\nDone: {args.output}")

    # ASCII summary
    ttr = _ttr_score(df_c)
    print("\n" + "=" * 78)
    print(f"{'Model':<22} {'TTR-Score':>10} {'LB':>6} {'CUAD':>6} {'IFEval':>7} {'MMLU':>6}")
    print("-" * 78)
    for model in _sort_models(ttr.to_frame()).index:
        row = {}
        for bm in BENCHMARKS:
            sub = df_c[(df_c["model"] == model) & (df_c["benchmark"] == bm)]
            row[bm] = sub["is_correct"].mean() * 100 if not sub.empty else float("nan")
        print(f"{model:<22} {ttr[model]:>9.1f}%"
              f" {row['legalbench']:>5.1f}% {row['cuad']:>5.1f}%"
              f" {row['ifeval']:>6.1f}% {row['mmlupro']:>5.1f}%")
    if partial:
        print("-" * 78)
        for m, b in partial:
            print(f"{m:<22} (partial: {', '.join(b)})")
    print("=" * 78)


if __name__ == "__main__":
    main()
